# -*- coding: utf-8 -*-
"""
qa-report — 차수별 통합테스트 결과서 결정론적 렌더러.

입력: JSON 페이로드 파일 경로 (argv[1]). 모든 집계는 호출측(커맨드 + Jira fetch)이
계산해 페이로드에 담아 전달한다. 본 스크립트는 순수 렌더러 + 메타 블록 기록만 수행
(셀 양식·색·테두리 고정). 양식은 코드 내장이라 외부 템플릿 파일 의존이 없다.

payload 스키마:
{
  "solution": "MYAPP",
  "project_key": "MYAPP",
  "round": 2,
  "period_start": "2026-06-01", "period_end": "2026-06-02",
  "build_ver": "1.2.0",
  "qa_staff": "QA: 최명인",
  "cause_field_available": true,
  "issues": [{"key","kind"(신규|이월),"cause","summary"}, ...],
  "agg": {"new":2,"carry":1,"cumulative_found":5,"resolved":2,"remaining":3},
  "cause_breakdown": {"배포/형상 관리": 1, ...},          # 원인별 건수(분포%는 렌더러가 계산)
  "trend": [{"round_label":"1차 통합테스트","new":3}, ...],  # 누적은 렌더러가 new running sum으로 계산
  "verdict": "판정: ...\n근거: ...",
  "output_path": "<출력폴더>/MYAPP QA 2차 통합테스트 결과서.xlsx",
  "meta": { ... 다음 차수가 읽을 동결값 ... }
}
사용: python build_report.py <payload.json>
"""
import sys, os, json

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    sys.stderr.write("openpyxl 필요: pip install openpyxl\n")
    sys.exit(2)

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

# ---- 양식 토큰 (코드 내장 고정 RGB) ----
FONT = "맑은 고딕"
C_TITLE = "FF2E76B6"   # theme8(accent5 5B9BD5) tint -0.25 해석값
C_SECTION = "FFF2F2F2"  # theme0(white) tint -0.05
C_TH = "FFB7DEE8"
C_TOTAL = "FFD9D9D9"
C_CARRY = "FFFCE4D6"
C_BLACK = "FF000000"
C_WHITE = "FFFFFFFF"

FILL_TITLE = PatternFill("solid", fgColor=C_TITLE)
FILL_SECTION = PatternFill("solid", fgColor=C_SECTION)
FILL_TH = PatternFill("solid", fgColor=C_TH)
FILL_TOTAL = PatternFill("solid", fgColor=C_TOTAL)
FILL_CARRY = PatternFill("solid", fgColor=C_CARRY)
NOFILL = PatternFill(fill_type=None)

F_TITLE = Font(name=FONT, sz=18, b=True, color=C_WHITE)
F_SECT = Font(name=FONT, sz=11, b=True, color=C_BLACK)
F_SUB = Font(name=FONT, sz=10, b=True, color=C_BLACK)
F_TH = Font(name=FONT, sz=10, b=True, color=C_BLACK)
F_THS = Font(name=FONT, sz=9, b=True, color=C_BLACK)
F_BOLD = Font(name=FONT, sz=10, b=True, color=C_BLACK)
F_DATA = Font(name=FONT, sz=10, b=False, color=C_BLACK)
F_DATAS = Font(name=FONT, sz=9, b=False, color=C_BLACK)

A_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
A_R = Alignment(horizontal="right", vertical="center", wrap_text=True)
A_TL = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN = Side("thin", color=C_BLACK)
MED = Side("medium", color=C_BLACK)

LAST = 9
LASTL = "I"

CAUSE_ORDER = [
    "배포/형상 관리", "기획/명세 누락", "사이드 이펙트 (회귀 오류)", "단순 구현/로직",
    "데이터 정합성/환경", "단위 테스트 누락", "기타(판별 불가)", "개선",
]


def build(payload):
    import datetime
    wb = openpyxl.Workbook()
    # 결정성: 생성/수정 시각을 고정해 동일 입력 → 바이트 동일 보장
    fixed = datetime.datetime(2020, 1, 1, 0, 0, 0)
    wb.properties.created = fixed
    wb.properties.modified = fixed
    ws = wb.active
    ws.title = "Report"
    ws.sheet_view.showGridLines = False
    for col, wdt in {"A": 1.6, "B": 15, "C": 9, "D": 11, "E": 11, "F": 11, "G": 9, "H": 9, "I": 13}.items():
        ws.column_dimensions[col].width = wdt

    def W(coord, val, font, fill, align, numfmt=None):
        c = ws[coord]
        c.value = val
        c.font = font
        c.fill = fill
        c.alignment = align
        if numfmt:
            c.number_format = numfmt
        return c

    def merge(rng):
        ws.merge_cells(rng)

    def grp(sc, ec, row, val, font, fill, align, numfmt=None):
        W(f"{sc}{row}", val, font, fill, align, numfmt)
        if sc != ec:
            merge(f"{sc}{row}:{ec}{row}")
            for ci in range(column_index_from_string(sc) + 1, column_index_from_string(ec) + 1):
                cc = ws.cell(row=row, column=ci)
                cc.font = font
                cc.fill = fill
                cc.alignment = align

    def rect(top, bottom, c1=2, c2=LAST):
        for r in range(top, bottom + 1):
            for ci in range(c1, c2 + 1):
                L = MED if ci == c1 else THIN
                R = MED if ci == c2 else THIN
                T = MED if r == top else THIN
                B = MED if r == bottom else THIN
                ws.cell(row=r, column=ci).border = Border(left=L, right=R, top=T, bottom=B)

    def section(row, text):
        W(f"B{row}", text, F_SECT, FILL_SECTION, A_L)
        merge(f"B{row}:{LASTL}{row}")
        for ci in range(3, LAST + 1):
            ws.cell(row=row, column=ci).fill = FILL_SECTION
        ws.row_dimensions[row].height = 18

    def subhdr(row, text):
        W(f"B{row}", text, F_SUB, NOFILL, A_L)
        merge(f"B{row}:{LASTL}{row}")
        rect(row, row)
        ws.row_dimensions[row].height = 18

    sol = payload["solution"]
    rnd = payload["round"]
    cause_on = payload.get("cause_field_available", True)
    agg = payload["agg"]

    # 제목
    W("B2", f"{sol} {rnd}차 통합테스트 결과 보고서", F_TITLE, FILL_TITLE, A_C)
    merge(f"B2:{LASTL}2")
    for ci in range(3, LAST + 1):
        ws.cell(row=2, column=ci).fill = FILL_TITLE
    ws.row_dimensions[2].height = 32

    # Test Status
    section(4, "Test Status")
    meta_rows = [
        ("QA 담당자", " " + payload.get("qa_staff", "")),
        ("QA 일정", f" {payload['period_start']} ~ {payload['period_end']}"),
        ("Build Ver.", " " + str(payload.get("build_ver", ""))),
    ]
    r = 6
    for lab, val in meta_rows:
        W(f"B{r}", lab, F_THS, FILL_TH, A_C)
        grp("C", LASTL, r, val, F_DATAS, NOFILL, A_L)
        ws.row_dimensions[r].height = 22
        r += 1
    rect(6, 8)

    # 상세 이슈
    section(10, "상세 이슈")
    if cause_on:
        W("B12", "BTS", F_THS, FILL_TH, A_C)
        W("C12", "구분", F_THS, FILL_TH, A_C)
        grp("D", "E", 12, "결함 원인 유형", F_THS, FILL_TH, A_C)
        grp("F", LASTL, 12, "상세 내용", F_THS, FILL_TH, A_C)
    else:
        W("B12", "BTS", F_THS, FILL_TH, A_C)
        W("C12", "구분", F_THS, FILL_TH, A_C)
        grp("D", LASTL, 12, "상세 내용", F_THS, FILL_TH, A_C)
    r = 13
    issues = payload.get("issues", [])
    for it in issues:
        fill = FILL_CARRY if it.get("kind") == "이월" else NOFILL
        W(f"B{r}", it.get("key", ""), F_DATAS, fill, A_C)
        W(f"C{r}", it.get("kind", ""), F_DATAS, fill, A_C)
        if cause_on:
            grp("D", "E", r, it.get("cause", ""), F_DATAS, fill, A_C)
            grp("F", LASTL, r, it.get("summary", ""), F_DATAS, fill, A_L)
        else:
            grp("D", LASTL, r, it.get("summary", ""), F_DATAS, fill, A_L)
        ws.row_dimensions[r].height = 20
        r += 1
    if issues:
        rect(12, 12 + len(issues))
    else:
        rect(12, 12)

    # Bug Status
    base = 12 + max(len(issues), 1) + 2
    section(base, "Bug Status")
    R = base + 2

    # 1) 이슈 수 집계
    subhdr(R, "1) 이슈 수 집계")
    R += 1
    grp("B", "E", R, "항목", F_TH, FILL_TH, A_C)
    grp("F", LASTL, R, "값", F_TH, FILL_TH, A_C)
    ws.row_dimensions[R].height = 20
    R += 1
    rate = (agg["resolved"] / agg["cumulative_found"]) if agg["cumulative_found"] else 0
    agg_rows = [
        ("금차 신규", agg["new"], None, False),
        ("금차 이월", agg["carry"], None, False),
        ("누적 발견 (1~%d차)" % rnd, agg["cumulative_found"], None, False),
        ("해결 완료", agg["resolved"], None, False),
        ("잔여 미해결", agg["remaining"], None, True),
        ("해결률", rate, "0.0%", True),
    ]
    top1 = R
    for lab, val, nf, bold in agg_rows:
        f = F_BOLD if bold else F_DATA
        grp("B", "E", R, lab, f, NOFILL, A_L)
        grp("F", LASTL, R, val, f, NOFILL, A_R, nf)
        ws.row_dimensions[R].height = 20
        R += 1
    rect(top1, R - 1)
    R += 1

    # 2) 결함 원인 유형별 분포 (필드 있을 때만)
    if cause_on:
        subhdr(R, "2) 결함 원인 유형별 분포")
        R += 1
        grp("B", "E", R, "결함 원인 유형", F_TH, FILL_TH, A_C)
        grp("F", "G", R, "건수", F_TH, FILL_TH, A_C)
        grp("H", LASTL, R, "비율", F_TH, FILL_TH, A_C)
        ws.row_dimensions[R].height = 20
        R += 1
        cb = payload.get("cause_breakdown", {})
        def _cnt(v):
            return v.get("count", 0) if isinstance(v, dict) else (v or 0)
        tot_c = sum(_cnt(cb.get(n, 0)) for n in CAUSE_ORDER)
        top2 = R
        for name in CAUSE_ORDER:
            cnt = _cnt(cb.get(name, 0))
            pct = (cnt / tot_c) if tot_c else 0
            grp("B", "E", R, name, F_DATA, NOFILL, A_L)
            grp("F", "G", R, cnt, F_DATA, NOFILL, A_R)
            grp("H", LASTL, R, pct, F_DATA, NOFILL, A_R, "0.0%")
            ws.row_dimensions[R].height = 19
            R += 1
        grp("B", "E", R, "Total", F_BOLD, FILL_TOTAL, A_C)
        grp("F", "G", R, tot_c, F_BOLD, FILL_TOTAL, A_R)
        grp("H", LASTL, R, (1 if tot_c else 0), F_BOLD, FILL_TOTAL, A_R, "0.0%")
        ws.row_dimensions[R].height = 20
        rect(top2 - 1, R)
        R += 2

    # 3) 차수별 추이
    subhdr(R, "3) 차수별 추이")
    R += 1
    grp("B", "E", R, "차수", F_TH, FILL_TH, A_C)
    grp("F", "G", R, "신규 발견", F_TH, FILL_TH, A_C)
    grp("H", LASTL, R, "누적 발견", F_TH, FILL_TH, A_C)
    ws.row_dimensions[R].height = 20
    R += 1
    top3 = R
    run = 0
    for t in payload.get("trend", []):
        new = t.get("new", 0)
        run += new  # 누적 = 신규 발견의 running sum (cumulative는 넘겨도 무시 — 불일치 방지)
        grp("B", "E", R, t["round_label"], F_BOLD, NOFILL, A_C)
        grp("F", "G", R, new, F_DATA, NOFILL, A_R)
        grp("H", LASTL, R, run, F_DATA, NOFILL, A_R)
        ws.row_dimensions[R].height = 20
        R += 1
    rect(top3 - 1, R - 1)
    R += 1

    # 결론 / 판정
    section(R, "결론 / 판정")
    R += 2
    W(f"B{R}", payload.get("verdict", ""), F_DATA, NOFILL, A_TL)
    merge(f"B{R}:{LASTL}{R + 4}")
    rect(R, R + 4)

    # 자기기술 메타 블록 (숨김 시트)
    mw = wb.create_sheet("_meta")
    mw.sheet_state = "hidden"
    mw["A1"] = "qa-report meta (자동 생성 — 수정 금지)"
    mw["A2"] = json.dumps(payload.get("meta", {}), ensure_ascii=False)

    out = payload["output_path"]
    os.makedirs(os.path.dirname(out) or ".", exist_ok=True)
    wb.save(out)
    return out


def main():
    if len(sys.argv) < 2:
        sys.stderr.write("usage: python build_report.py <payload.json>\n")
        sys.exit(1)
    with open(sys.argv[1], encoding="utf-8") as f:
        payload = json.load(f)
    out = build(payload)
    print("SAVED:", out)


if __name__ == "__main__":
    main()
